import base64
import pandas as pd
import traceback
from celery import shared_task
from datetime import datetime
from django.utils.timezone import make_aware
from io import BytesIO
from openpyxl import load_workbook

from investment_tracker.accessors import AssetsAccessor, AssetClassesAccessor, CountriesAccessor
from investment_tracker.models import AssetsModel, TransactionsModel
from investment_tracker.services import AsyncTasksService
from investment_tracker.utils.transactions_utils import to_lower_denomination

INDMONEY_TABLE_COLS = {
    "SOURCE_HOLDING_ID": "Source holding ID",
    "TRADE_DATE": "Trade Date",
    "INVESTMENT_NAME": "Investment name",
    "ASSET_TYPE": "Asset Type",
    "BUY_UNITS": "Buy units",
    "SELL_UNITS": "Sell units",
    "DIVIDEND_REINVESTED_UNITS": "Dividend reinvested units",
    "CASH_INFLOW": "Cash inflow",
    "CASH_OUTFLOW": "Cash outflow",
    "BENCHMARK_CASH_INFLOW": "Benchmark cash inflow",
    "BENCHMARK_CASH_OUTFLOW": "Benchmark cash outflow",
    "DIVIDEND_AMOUNT": "Dividend Amount",
}


def extract(source_data: str) -> list[dict]:
    """Extracts data from INDMoney Tradebook xlsx"""
    _, file_data = source_data.split(";base64,")
    data = BytesIO(base64.b64decode(file_data))
    wb = load_workbook(filename=data, read_only=True)
    data = []
    for sheetname in wb.sheetnames:
        found_cols = False
        found_data = False
        sheet = wb[sheetname]
        sheet_transactions = []
        for row in sheet.values:
            if found_data and row[0] is None:
                break
            elif found_data:
                sheet_transactions.append(row)
            elif found_cols and not found_data:
                found_data = True
            elif list(row) == list(INDMONEY_TABLE_COLS.values()):
                found_cols = True
        sheet_transactions_df = pd.DataFrame(sheet_transactions, columns=list(INDMONEY_TABLE_COLS.values()))
        data.extend(sheet_transactions_df.to_dict("records"))
    return data


def transform(extracted_data: list[dict]) -> dict:
    """Transforms extracted data from INDMoney Tradebook xlsx"""
    asset_classes = AssetClassesAccessor().get_asset_classes()
    asset_classes_map = {asset_class.name: asset_class.id for asset_class in asset_classes}
    countries = CountriesAccessor().get_countries()
    countries_map = {country.code: country.id for country in countries}
    transactions_df = pd.DataFrame(extracted_data, columns=list(INDMONEY_TABLE_COLS.values()))
    # Ignoring Dividends
    transactions_df = transactions_df[
        (transactions_df[INDMONEY_TABLE_COLS["BUY_UNITS"]] != "0")
        | (transactions_df[INDMONEY_TABLE_COLS["SELL_UNITS"]] != "0")
    ]
    names = set(transactions_df[INDMONEY_TABLE_COLS["INVESTMENT_NAME"]].tolist())
    existing_assets = AssetsAccessor().get_assets(tickers=list(names) + ["USD"])
    assets_map = {asset.ticker: asset for asset in existing_assets}
    missing_assets = names.difference(assets_map.keys())
    asset_type_map = transactions_df.set_index(INDMONEY_TABLE_COLS["INVESTMENT_NAME"]).to_dict()[
        INDMONEY_TABLE_COLS["ASSET_TYPE"]
    ]
    new_assets = []
    for ticker in missing_assets:
        asset_type = asset_type_map[ticker]
        asset_class = 0
        if asset_type == "US_STOCK":
            asset_class = asset_classes_map["Stock"]
        elif asset_type == "MF":
            asset_class = asset_classes_map["Mutual Fund"]
        asset = AssetsModel(name=ticker, ticker=ticker, asset_class_id=asset_class, country_id=countries_map["USA"])
        assets_map[ticker] = asset
        new_assets.append(asset)
    transactions_df["supply_asset"] = transactions_df.apply(
        lambda row: assets_map["USD"]
        if row[INDMONEY_TABLE_COLS["BUY_UNITS"]] != "0"
        else assets_map[row[INDMONEY_TABLE_COLS["INVESTMENT_NAME"]]],
        axis=1,
    )
    transactions_df["supply_value"] = transactions_df.apply(
        lambda row: to_lower_denomination(row[INDMONEY_TABLE_COLS["CASH_INFLOW"]], asset=assets_map["USD"])
        if row[INDMONEY_TABLE_COLS["BUY_UNITS"]] != "0"
        else to_lower_denomination(
            row[INDMONEY_TABLE_COLS["SELL_UNITS"]], asset=assets_map[row[INDMONEY_TABLE_COLS["INVESTMENT_NAME"]]]
        ),
        axis=1,
    )
    transactions_df["receive_asset"] = transactions_df.apply(
        lambda row: assets_map[row[INDMONEY_TABLE_COLS["INVESTMENT_NAME"]]]
        if row[INDMONEY_TABLE_COLS["BUY_UNITS"]] != "0"
        else assets_map["USD"],
        axis=1,
    )
    transactions_df["receive_value"] = transactions_df.apply(
        lambda row: to_lower_denomination(
            row[INDMONEY_TABLE_COLS["BUY_UNITS"]], asset=assets_map[row[INDMONEY_TABLE_COLS["INVESTMENT_NAME"]]]
        )
        if row[INDMONEY_TABLE_COLS["BUY_UNITS"]] != "0"
        else to_lower_denomination(row[INDMONEY_TABLE_COLS["CASH_OUTFLOW"]], asset=assets_map["USD"]) * -1,
        axis=1,
    )
    transactions_df["transacted_at"] = transactions_df.apply(
        lambda row: make_aware(datetime.strptime(row[INDMONEY_TABLE_COLS["TRADE_DATE"]], "%Y-%m-%d")),
        axis=1,
    )
    transactions_df = transactions_df.drop(columns=list(INDMONEY_TABLE_COLS.values()))
    transactions = transactions_df.to_dict("records")
    transactions = [TransactionsModel(**transaction) for transaction in transactions]
    return {"transactions": transactions, "assets": new_assets}


def load(transformed_data: dict) -> None:
    """Loads transactions to database"""
    assets = transformed_data.get("assets", [])
    AssetsModel.objects.bulk_create(assets)
    transactions = transformed_data.get("transactions", [])
    TransactionsModel.objects.bulk_create(transactions)


@shared_task
def run(async_task_id: int, source_data: str) -> None:
    try:
        AsyncTasksService().set_in_progress(async_task_id, percentage=25)
        extracted_data = extract(source_data)
        AsyncTasksService().update_progress(async_task_id, 50)
        transformed_data = transform(extracted_data)
        AsyncTasksService().update_progress(async_task_id, 75)
        load(transformed_data)
        AsyncTasksService().set_completed(async_task_id)
    except Exception as ex:
        traceback.print_exc()
        AsyncTasksService().set_failed(async_task_id)
