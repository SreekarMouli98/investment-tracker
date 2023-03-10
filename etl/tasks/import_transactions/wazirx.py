import logging
from datetime import datetime
from io import BytesIO

import pandas as pd
import pydash
from django.utils.timezone import make_aware
from openpyxl import load_workbook

from etl.tasks.base_etl import ETL
from etl.tasks.import_transactions.mixins.load_mixin import LoadMixin
from etl.utils.common_utils import decode_base64_data
from investment_tracker.accessors import AssetClassesAccessor
from investment_tracker.accessors import AssetsAccessor
from investment_tracker.models import AssetsModel
from investment_tracker.models import TransactionsModel
from investment_tracker.services.conversion_rates_services import ConversionRatesService
from investment_tracker.utils.transactions_utils import get_base_asset
from investment_tracker.utils.transactions_utils import to_lower_denomination

logger = logging.getLogger(__name__)

EXCHANGE_TRADES_TABLE_COLS = {
    "DATE": "Date",
    "MARKET": "Market",
    "PRICE": "Price",
    "VOLUME": "Volume",
    "TOTAL_PRICE_X_VOLUME": "Total (Price x Volume)",
    "TRADE_TYPE": "Trade Type",
    "FEE_PAID_IN": "Fee Paid in",
    "FEE_AMOUNT": "Fee Amount",
    "TDS_PAID_IN": "TDS Paid in",
    "TDS_AMOUNT": "TDS Amount",
}

P2P_TRADES_TABLE_COLS = {
    "DATE": "Date",
    "MARKET": "Market",
    "PRICE": "Price",
    "VOLUME": "Volume",
    "TOTAL_PRICE_X_VOLUME": "Total (Price x Volume)",
    "TRADE_TYPE": "Trade Type",
    "TDS_PAID_IN": "TDS Paid in",
    "TDS_AMOUNT": "TDS Amount",
}


class ImportTransactionsFromWazirxETL(LoadMixin, ETL):
    def extract(self, source_data: str) -> list[dict]:
        """Extracts data from Wazirx Tradebook xlsx"""
        logger.info("[Import Transactions ETL]: Wazirx -> Extract -> Begin")
        data = BytesIO(decode_base64_data(source_data))
        wb = load_workbook(filename=data, read_only=True)
        exchange_trades_sheet = wb["Exchange Trades"]
        i = 2
        rows = []
        while True:
            if exchange_trades_sheet.cell(row=i, column=1).value == None:
                break
            row = []
            for j in range(len(EXCHANGE_TRADES_TABLE_COLS.keys())):
                row.append(exchange_trades_sheet.cell(row=i, column=j + 1).value)
            rows.append(row)
            i += 1
        exchange_trades_df = pd.DataFrame(
            rows, columns=list(EXCHANGE_TRADES_TABLE_COLS.values())
        )
        p2p_trades_sheet = wb["P2P Trades"]
        i = 2
        rows = []
        while True:
            if p2p_trades_sheet.cell(row=i, column=1).value == None:
                break
            row = []
            for j in range(len(P2P_TRADES_TABLE_COLS.keys())):
                row.append(p2p_trades_sheet.cell(row=i, column=j + 1).value)
            rows.append(row)
            i += 1
        p2p_trades_df = pd.DataFrame(rows, columns=list(P2P_TRADES_TABLE_COLS.values()))
        data = {
            "exchange_trades": exchange_trades_df.to_dict("records"),
            "p2p_trades": p2p_trades_df.to_dict("records"),
        }
        logger.info("[Import Transactions ETL]: Wazirx -> Extract -> End")
        return data

    def transform(self, extracted_data: list[dict]) -> dict:
        """Transforms extracted data from Wazirx Tradebook xlsx"""
        logger.info("[Import Transactions ETL]: Wazirx -> Transform -> Begin")
        asset_classes = AssetClassesAccessor().get_asset_classes()
        asset_classes_map = {
            asset_class.name: asset_class.id for asset_class in asset_classes
        }
        exchange_trades_df = pd.DataFrame(
            extracted_data["exchange_trades"],
            columns=list(EXCHANGE_TRADES_TABLE_COLS.values()),
        )
        p2p_trades_df = pd.DataFrame(
            extracted_data["p2p_trades"], columns=list(P2P_TRADES_TABLE_COLS.values())
        )
        p2p_trades_df[EXCHANGE_TRADES_TABLE_COLS["FEE_PAID_IN"]] = "INR"
        p2p_trades_df[EXCHANGE_TRADES_TABLE_COLS["FEE_AMOUNT"]] = ""
        data = exchange_trades_df.to_dict("records")
        data.extend(p2p_trades_df.to_dict("records"))
        transactions_df = pd.DataFrame(
            data, columns=list(EXCHANGE_TRADES_TABLE_COLS.values())
        )
        transactions_df["supply_asset"] = transactions_df[
            EXCHANGE_TRADES_TABLE_COLS["FEE_PAID_IN"]
        ]
        transactions_df["receive_asset"] = transactions_df.apply(
            lambda row: row[EXCHANGE_TRADES_TABLE_COLS["MARKET"]].replace(
                row["supply_asset"], ""
            ),
            axis=1,
        )
        tickers = set(
            transactions_df["supply_asset"].tolist()
            + transactions_df["receive_asset"].tolist()
        )
        existing_assets = AssetsAccessor().get_assets(tickers=list(tickers) + ["INR"])
        assets_map = {asset.ticker: asset for asset in existing_assets}
        missing_assets = tickers.difference(assets_map.keys())
        new_assets = []
        for ticker in missing_assets:
            asset = AssetsModel(
                name=ticker, ticker=ticker, asset_class_id=asset_classes_map["Crypto"]
            )
            assets_map[ticker] = asset
            new_assets.append(asset)
        base_asset = get_base_asset()
        transactions_df["supply_asset"] = transactions_df.apply(
            lambda row: assets_map[row["supply_asset"]], axis=1
        )
        transactions_df["supply_value"] = transactions_df.apply(
            lambda row: to_lower_denomination(
                row[EXCHANGE_TRADES_TABLE_COLS["TOTAL_PRICE_X_VOLUME"]],
                asset=row["supply_asset"],
            ),
            axis=1,
        )
        transactions_df["receive_asset"] = transactions_df.apply(
            lambda row: assets_map[row["receive_asset"]], axis=1
        )
        transactions_df["receive_value"] = transactions_df.apply(
            lambda row: to_lower_denomination(
                row[EXCHANGE_TRADES_TABLE_COLS["VOLUME"]], asset=row["receive_asset"]
            ),
            axis=1,
        )
        transactions_df["transacted_at"] = transactions_df.apply(
            lambda row: make_aware(
                datetime.strptime(
                    row[EXCHANGE_TRADES_TABLE_COLS["DATE"]], "%Y-%m-%d %H:%M:%S"
                )
            ),
            axis=1,
        )
        conversion_rates = []
        failed_conversions = []
        cached_conversion_rates = {}
        transactions_df["supply_base_conv_rate"] = transactions_df.apply(
            lambda row: to_lower_denomination(
                row[EXCHANGE_TRADES_TABLE_COLS["PRICE"]], asset=base_asset
            )
            if row["receive_asset"] == base_asset
            else ConversionRatesService().get_conversion_rate_cached(
                row["supply_asset"],
                base_asset,
                date=row["transacted_at"],
                cached_conversion_rates=cached_conversion_rates,
                skip_persist=True,
                new_conversion_rates=conversion_rates,
                failed_conversions=failed_conversions,
            ),
            axis=1,
        )
        transactions_df["receive_base_conv_rate"] = transactions_df.apply(
            lambda row: to_lower_denomination(1, asset=base_asset)
            if row["receive_asset"] == base_asset
            else ConversionRatesService().get_conversion_rate_cached(
                row["receive_asset"],
                base_asset,
                date=row["transacted_at"],
                cached_conversion_rates=cached_conversion_rates,
                skip_persist=True,
                new_conversion_rates=conversion_rates,
                failed_conversions=failed_conversions,
            ),
            axis=1,
        )
        failed_conversions = [
            f"{from_asset.ticker}/{to_asset.ticker}"
            for from_asset, to_asset in failed_conversions
        ]
        failed_conversions = pydash.uniq(failed_conversions)
        transactions_df = transactions_df.drop(
            columns=list(EXCHANGE_TRADES_TABLE_COLS.values())
        )
        transactions = transactions_df.to_dict("records")
        transactions = [
            TransactionsModel(**transaction) for transaction in transactions
        ]
        logger.info("[Import Transactions ETL]: Wazirx -> Transform -> End")
        return {
            "transactions": transactions,
            "assets": new_assets,
            "conversion_rates": conversion_rates,
            "failed_conversions": failed_conversions,
        }
