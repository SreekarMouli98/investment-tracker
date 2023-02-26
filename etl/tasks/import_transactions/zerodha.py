import pandas as pd
from datetime import datetime
from django.utils.timezone import make_aware
from io import BytesIO
from openpyxl import load_workbook

from etl.tasks.base_etl import ETL
from etl.tasks.import_transactions.mixins.load_mixin import LoadMixin
from etl.utils.common_utils import decode_base64_data
from investment_tracker.accessors import AssetsAccessor, AssetClassesAccessor, CountriesAccessor
from investment_tracker.models import AssetsModel, TransactionsModel
from investment_tracker.services.conversion_rates_services import ConversionRatesService
from investment_tracker.utils.transactions_utils import (
    get_base_asset,
    to_higher_denomination,
    to_lower_denomination,
)

ZERODHA_TABLE_COLS = {
    "ORDER_NO": "Order No.",
    "ORDER_TIME": "Order Time.",
    "TRADE_NO": "Trade No.",
    "TRADE_TIME": "Trade Time",
    "SECURITY_CONTRACT_DESCRIPTION": "Security/Contract Description",
    "BUY_B_SELL_S": "Buy(B) / Sell(S)",
    "QUANTITY": "Quantity",
    "GROSS_RATE_TRADE_PRICE_PER_UNIT_RS": "Gross Rate / Trade Price Per Unit (Rs)",
    "BROKERAGE_PER_UNIT_RS": "Brokerage per Unit (Rs)",
    "NET_RATE_PER_UNIT_RS": "Net Rate per Unit (Rs)",
    "CLOSING_RATE_PER_UNIT_ONLY_FOR_DERIVATIVES_RS": "Closing Rate per Unit (Only for Derivatives) (Rs)",
    "NET_TOTAL_BEFORE_LEVIES_RS": "Net Total (Before Levies) (Rs)",
    "REMARKS": "Remarks",
    "ISIN": "ISIN",
    "DATE": "Date",
}


class ImportTransactionsFromZerodhaETL(LoadMixin, ETL):
    def extract(self, source_data: str) -> list[dict]:
        """Extracts data from Zerodha Tradebook xlsx"""
        data = BytesIO(decode_base64_data(source_data))
        wb = load_workbook(filename=data, read_only=True)
        data = []
        for sheetname in wb.sheetnames:
            found_cols = False
            found_data = False
            sheet = wb[sheetname]
            sheet_transactions = []
            for row in sheet.values:
                if found_data and row[0] is None:
                    break
                elif found_data:
                    sheet_transactions.append(row)
                elif found_cols and not found_data:
                    found_data = True
                elif list(row) == list(ZERODHA_TABLE_COLS.values())[:-1]:
                    found_cols = True
            sheet_transactions_df = pd.DataFrame(sheet_transactions, columns=list(ZERODHA_TABLE_COLS.values())[:-1])
            sheet_transactions_df[ZERODHA_TABLE_COLS["DATE"]] = sheetname
            data.extend(sheet_transactions_df.to_dict("records"))
        return data

    def transform(self, extracted_data: list[dict]) -> dict:
        """Transforms extracted data from Zerodha Tradebook xlsx"""
        asset_classes = AssetClassesAccessor().get_asset_classes()
        asset_classes_map = {asset_class.name: asset_class.id for asset_class in asset_classes}
        countries = CountriesAccessor().get_countries()
        countries_map = {country.code: country.id for country in countries}
        transactions_df = pd.DataFrame(extracted_data, columns=list(ZERODHA_TABLE_COLS.values()))
        transactions_df[ZERODHA_TABLE_COLS["SECURITY_CONTRACT_DESCRIPTION"]] = transactions_df[
            ZERODHA_TABLE_COLS["SECURITY_CONTRACT_DESCRIPTION"]
        ].apply(lambda item: item.split()[0])
        tickers = set(transactions_df[ZERODHA_TABLE_COLS["SECURITY_CONTRACT_DESCRIPTION"]].tolist())
        existing_assets = AssetsAccessor().get_assets(tickers=list(tickers) + ["INR"])
        assets_map = {asset.ticker: asset for asset in existing_assets}
        missing_assets = tickers.difference(assets_map.keys())
        new_assets = []
        for ticker in missing_assets:
            asset = AssetsModel(
                name=ticker, ticker=ticker, asset_class_id=asset_classes_map["Stock"], country_id=countries_map["IND"]
            )
            assets_map[ticker] = asset
            new_assets.append(asset)
        base_asset = get_base_asset()
        transactions_df["supply_asset"] = transactions_df.apply(
            lambda row: assets_map["INR"]
            if row[ZERODHA_TABLE_COLS["BUY_B_SELL_S"]] in ("buy", "B")
            else assets_map[row[ZERODHA_TABLE_COLS["SECURITY_CONTRACT_DESCRIPTION"]]],
            axis=1,
        )
        transactions_df["supply_value"] = transactions_df.apply(
            lambda row: to_lower_denomination(
                row[ZERODHA_TABLE_COLS["NET_TOTAL_BEFORE_LEVIES_RS"]], asset=assets_map["INR"]
            )
            * -1
            if row[ZERODHA_TABLE_COLS["BUY_B_SELL_S"]] in ("buy", "B")
            else to_lower_denomination(
                row[ZERODHA_TABLE_COLS["QUANTITY"]],
                asset=assets_map[row[ZERODHA_TABLE_COLS["SECURITY_CONTRACT_DESCRIPTION"]]],
            ),
            axis=1,
        )
        transactions_df["receive_asset"] = transactions_df.apply(
            lambda row: assets_map[row[ZERODHA_TABLE_COLS["SECURITY_CONTRACT_DESCRIPTION"]]]
            if row[ZERODHA_TABLE_COLS["BUY_B_SELL_S"]] in ("buy", "B")
            else assets_map["INR"],
            axis=1,
        )
        transactions_df["receive_value"] = transactions_df.apply(
            lambda row: to_lower_denomination(
                row[ZERODHA_TABLE_COLS["QUANTITY"]],
                asset=assets_map[row[ZERODHA_TABLE_COLS["SECURITY_CONTRACT_DESCRIPTION"]]],
            )
            if row[ZERODHA_TABLE_COLS["BUY_B_SELL_S"]] in ("buy", "B")
            else to_lower_denomination(row[ZERODHA_TABLE_COLS["NET_TOTAL_BEFORE_LEVIES_RS"]], asset=assets_map["INR"]),
            axis=1,
        )
        transactions_df["transacted_at"] = transactions_df.apply(
            lambda row: make_aware(datetime.strptime(row[ZERODHA_TABLE_COLS["DATE"]], "%d-%m-%Y")),
            axis=1,
        )
        transactions_df["supply_base_conv_rate"] = transactions_df.apply(
            lambda row: ConversionRatesService().get_conversion_rate(
                assets_map["INR"], base_asset, date=row["transacted_at"]
            )
            if row[ZERODHA_TABLE_COLS["BUY_B_SELL_S"]] in ("buy", "B")
            else (
                to_higher_denomination(row["receive_value"], asset_class_instance=assets_map["INR"].asset_class)
                / to_higher_denomination(
                    row["supply_value"],
                    asset_class_instance=assets_map[
                        row[ZERODHA_TABLE_COLS["SECURITY_CONTRACT_DESCRIPTION"]]
                    ].asset_class,
                )
            )
            * ConversionRatesService().get_conversion_rate(assets_map["INR"], base_asset, date=row["transacted_at"]),
            axis=1,
        )
        transactions_df["receive_base_conv_rate"] = transactions_df.apply(
            lambda row: (
                to_higher_denomination(row["supply_value"], asset_class_instance=assets_map["INR"].asset_class)
                / to_higher_denomination(
                    row["receive_value"],
                    asset_class_instance=assets_map[
                        row[ZERODHA_TABLE_COLS["SECURITY_CONTRACT_DESCRIPTION"]]
                    ].asset_class,
                )
            )
            * ConversionRatesService().get_conversion_rate(assets_map["INR"], base_asset, date=row["transacted_at"])
            if row[ZERODHA_TABLE_COLS["BUY_B_SELL_S"]] in ("buy", "B")
            else ConversionRatesService().get_conversion_rate(assets_map["INR"], base_asset, date=row["transacted_at"]),
            axis=1,
        )
        transactions_df = transactions_df.drop(columns=list(ZERODHA_TABLE_COLS.values()))
        transactions = transactions_df.to_dict("records")
        transactions = [TransactionsModel(**transaction) for transaction in transactions]
        return {"transactions": transactions, "assets": new_assets}
